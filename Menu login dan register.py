import tkinter as tk
from tkinter import messagebox
from tkinter import Entry, Button, Frame, Label
from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook
import os
from tkinter import messagebox, ttk
import random

def register():
    username = entry_register_username.get()
    password = entry_register_password.get()

    if not username or not password:
        messagebox.showwarning("Input Error", "Username and Password cannot be empty")
        return

    dataakun = []
    file_path = 'dataakun.xlsx'

    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            dataakun.append({'username': row[0], 'password': row[1]})
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['username', 'password'])
        workbook.save(file_path)

    username_ada = False

    for akun in dataakun:
        if username == akun['username']:
            messagebox.showwarning("Error", "Username already exists! Choose another one.")
            username_ada = True
            break

    if not username_ada:
        databaru = {'username': username, 'password': password}
        sheet.append([username, password])
        workbook.save(file_path)
        messagebox.showinfo("Success", "Registration successful")

def login():
    username = entry_login_username.get()
    password = entry_login_password.get()

    if not username or not password:
        messagebox.showwarning("Input Error", "Username and Password cannot be empty")
        return

    dataakun = []
    file_path = 'dataakun.xlsx'

    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            dataakun.append({'username': row[0], 'password': row[1]})
    else:
        messagebox.showwarning("Error", "File dataakun.xlsx not found")
        return

    login_berhasil = False
    for akun in dataakun:
        if username == akun['username'] and password == akun['password']:
            messagebox.showinfo("Success", "Login successful")
            show_main(username)
            login_berhasil = True
            break

    if not login_berhasil:
        messagebox.showwarning("Error", "Account not found or incorrect password")

def show_register_frame():
    register_frame.place(x=550, y=360)
    login_frame.place_forget()
    main_frame.place_forget()

def show_login_frame():
    login_frame.place(x=550, y=360)
    register_frame.place_forget()
    main_frame.place_forget()

def show_main(username=None):
    root.withdraw()
    global window
    window = tk.Toplevel(root)
    window.title("Selamat Datang")
    window.geometry("1200x800")
    window.configure(bg="White")
    window.resizable(True, True)

    new_bg_image = Image.open(r"C:\prokommmm\Tugas-Besar\Sapa.png")
    new_bg_image = new_bg_image.resize((1200, 800), Image.Resampling.LANCZOS)
    new_photo = ImageTk.PhotoImage(new_bg_image)

    background_label = tk.Label(window, image=new_photo)
    background_label.place(x=0, y=0, relwidth=1, relheight=1)
    background_label.image = new_photo
    
    new_frame = Frame(window, width=350, height=100, bg="#354f00")
    new_frame.place(x=465, y=450)

    Button(new_frame, width=38, height=1, text="Lanjutkan", fg="white", bg="#57a1f8", command=halaman_utama).place(x=40, y=20)
    Button(new_frame, width=38, height=1, text="Logout", fg="white", bg="#57a1f8", command=back_login).place(x=40, y=50)

    def new_background(event=None):
        width = window.winfo_width()
        height = window.winfo_height()
        resized_image = new_bg_image.resize((width, height), Image.Resampling.LANCZOS)
        new_photo = ImageTk.PhotoImage(resized_image)
        background_label.config(image=new_photo)
        background_label.image = new_photo

    window.bind("<Configure>", new_background)
    window.mainloop()

def back_login():
    window.destroy()
    root.deiconify()

def update_background(event=None):
    new_width = root.winfo_width()
    new_height = root.winfo_height()
    resized_bg_image = bg_image.resize((new_width, new_height), Image.Resampling.LANCZOS)
    bg_photo = ImageTk.PhotoImage(resized_bg_image)
    background_label.config(image=bg_photo)
    background_label.image = bg_photo

def halaman_utama():
    window.withdraw()
    global window2
    window2 = tk.Toplevel(window)
    window2.title("Halaman Utama")
    window2.geometry("1200x800")
    window2.configure(bg="White")
    window2.resizable(True, True)

    backg_image = Image.open(r"C:\prokommmm\Tugas-Besar\Main menu baru banget.png")
    backg_image = backg_image.resize((1200, 800), Image.Resampling.LANCZOS)
    photo_baru = ImageTk.PhotoImage(backg_image)

    bg_label = tk.Label(window2, image=photo_baru)
    bg_label.place(x=0, y=0, relwidth=1, relheight=1)
    bg_label.image = photo_baru

    Button(window2, width=20, height=1, text="Memilih Wisata", fg="white", bg="#57a1f8", command=Memilih_Wisata).place(x=175, y=540)
    Button(window2, width=20, height=1, text="Reschedule Tiket", fg="white", bg="#57a1f8", command=Resche_Tiket).place(x=500, y=600)
    Button(window2, width=20, height=1, text="Refund Tiket", fg="white", bg="#57a1f8", command=Kembalikan_Tiket).place(x=840, y=540)
    Button(window2, width=20, height=1, text="Riwayat Pembelian", fg="white", bg="#57a1f8", command=Riwayat_Tiket).place(x=1180, y=600)

    def bg_baru(event=None):
        width_baru = window2.winfo_width()
        height_baru = window2.winfo_height()
        resized_image_baru = backg_image.resize((width_baru, height_baru), Image.Resampling.LANCZOS)
        photo_baru = ImageTk.PhotoImage(resized_image_baru)
        bg_label.config(image=photo_baru)
        bg_label.image = photo_baru

    window2.bind("<Configure>", bg_baru)
    window2.mainloop()

def back_Menu():
    window2.destroy()
    root.deiconify()

def update_background(event=None):
    new_width = root.winfo_width()
    new_height = root.winfo_height()
    resized_bg_image = bg_image.resize((new_width, new_height), Image.Resampling.LANCZOS)
    bg_photo = ImageTk.PhotoImage(resized_bg_image)
    background_label.config(image=bg_photo)
    background_label.image = bg_photo

def Memilih_Wisata():
    window2.withdraw()
    global window3
    window3 = tk.Toplevel(window2)
    window3.title("MEMILIH WISATA")
    window3.geometry("1200x800")
    window3.configure(bg="White")
    window3.resizable(True, True)

    BG_image = Image.open(r"C:\prokommmm\Tugas-Besar\memilih wisata.png")
    BG_image = BG_image.resize((1200, 800), Image.Resampling.LANCZOS)
    photo_BR = ImageTk.PhotoImage(BG_image)

    BG_label = tk.Label(window3, image=photo_BR)
    BG_label.place(x=0, y=0, relwidth=1, relheight=1)
    BG_label.image = photo_BR
    
    Solsaf = Frame(window3, width=1000, height=100, bg="#354f00")
    Solsaf.place(x=300, y=50)
    deskripsi1 = "Solo Safari adalah tempat wisata baru yang menarik di Solo, Jawa Tengah. Diresmikan pada tahun 2023, Solo Safari menawarkan pengalaman wisata alam dan edukasi dengan konsep kebun binatang modern."
    deskripsi2 = "Fasilitas yang didapat apabila berkunjung ke Solo Safari antara lain: Zona Hewan, Petting Zoo, Safari Ride, dan Edukasi Konservasi."
    deskripsi3 = "Dengan harga Rp 25.000,00/orang, Anda bisa merasakan petualangan seru serta menambah pengetahuan tentang dunia satwa. Dengan fasilitas lengkap dan pengalaman yang edukatif."

    label_deskripsi1 = Label(Solsaf, text=deskripsi1, bg="#354f00", fg="white")
    label_deskripsi1.pack(pady=10)
    label_deskripsi2 = Label(Solsaf, text=deskripsi2, bg="#354f00", fg="white")
    label_deskripsi2.pack(pady=3.5)
    label_deskripsi3 = Label(Solsaf, text=deskripsi3, bg="#354f00", fg="white")
    label_deskripsi3.pack(pady=3.5)

    Lokananta = Frame(window3, width=1000, height=100, bg="#354f00")
    Lokananta.place(x=300, y=190)
    deskripsi1 = "Solo Safari adalah tempat wisata baru yang menarik di Solo, Jawa Tengah. Diresmikan pada tahun 2023, Solo Safari menawarkan pengalaman wisata alam dan edukasi dengan konsep kebun binatang modern."
    deskripsi2 = "Fasilitas yang didapat apabila berkunjung ke Solo Safari antara lain: Zona Hewan, Petting Zoo, Safari Ride, dan Edukasi Konservasi."
    deskripsi3 = "Dengan harga Rp 25.000,00/orang, Anda bisa merasakan petualangan seru serta menambah pengetahuan tentang dunia satwa. Dengan fasilitas lengkap dan pengalaman yang edukatif."

    label_deskripsi1 = Label(Lokananta, text=deskripsi1, bg="#354f00", fg="white")
    label_deskripsi1.pack(pady=10)
    label_deskripsi2 = Label(Lokananta, text=deskripsi2, bg="#354f00", fg="white")
    label_deskripsi2.pack(pady=3.5)
    label_deskripsi3 = Label(Lokananta, text=deskripsi3, bg="#354f00", fg="white")
    label_deskripsi3.pack(pady=3.5)

    Balekambang = Frame(window3, width=1000, height=100, bg="#354f00")
    Balekambang.place(x=300, y=330)
    deskripsi1 = "Solo Safari adalah tempat wisata baru yang menarik di Solo, Jawa Tengah. Diresmikan pada tahun 2023, Solo Safari menawarkan pengalaman wisata alam dan edukasi dengan konsep kebun binatang modern."
    deskripsi2 = "Fasilitas yang didapat apabila berkunjung ke Solo Safari antara lain: Zona Hewan, Petting Zoo, Safari Ride, dan Edukasi Konservasi."
    deskripsi3 = "Dengan harga Rp 25.000,00/orang, Anda bisa merasakan petualangan seru serta menambah pengetahuan tentang dunia satwa. Dengan fasilitas lengkap dan pengalaman yang edukatif."

    label_deskripsi1 = Label(Balekambang, text=deskripsi1, bg="#354f00", fg="white")
    label_deskripsi1.pack(pady=10)
    label_deskripsi2 = Label(Balekambang, text=deskripsi2, bg="#354f00", fg="white")
    label_deskripsi2.pack(pady=3.5)
    label_deskripsi3 = Label(Balekambang, text=deskripsi3, bg="#354f00", fg="white")
    label_deskripsi3.pack(pady=3.5)

    Mangkunegaran = Frame(window3, width=1000, height=100, bg="#354f00")
    Mangkunegaran.place(x=300, y=470)
    deskripsi1 = "Solo Safari adalah tempat wisata baru yang menarik di Solo, Jawa Tengah. Diresmikan pada tahun 2023, Solo Safari menawarkan pengalaman wisata alam dan edukasi dengan konsep kebun binatang modern."
    deskripsi2 = "Fasilitas yang didapat apabila berkunjung ke Solo Safari antara lain: Zona Hewan, Petting Zoo, Safari Ride, dan Edukasi Konservasi."
    deskripsi3 = "Dengan harga Rp 25.000,00/orang, Anda bisa merasakan petualangan seru serta menambah pengetahuan tentang dunia satwa. Dengan fasilitas lengkap dan pengalaman yang edukatif."

    label_deskripsi1 = Label(Mangkunegaran, text=deskripsi1, bg="#354f00", fg="white")
    label_deskripsi1.pack(pady=10)
    label_deskripsi2 = Label(Mangkunegaran, text=deskripsi2, bg="#354f00", fg="white")
    label_deskripsi2.pack(pady=3.5)
    label_deskripsi3 = Label(Mangkunegaran, text=deskripsi3, bg="#354f00", fg="white")
    label_deskripsi3.pack(pady=3.5)

    Tumurun = Frame(window3, width=1000, height=100, bg="#354f00")
    Tumurun.place(x=300, y=620)
    deskripsi1 = "Solo Safari adalah tempat wisata baru yang menarik di Solo, Jawa Tengah. Diresmikan pada tahun 2023, Solo Safari menawarkan pengalaman wisata alam dan edukasi dengan konsep kebun binatang modern."
    deskripsi2 = "Fasilitas yang didapat apabila berkunjung ke Solo Safari antara lain: Zona Hewan, Petting Zoo, Safari Ride, dan Edukasi Konservasi."
    deskripsi3 = "Dengan harga Rp 25.000,00/orang, Anda bisa merasakan petualangan seru serta menambah pengetahuan tentang dunia satwa. Dengan fasilitas lengkap dan pengalaman yang edukatif."

    label_deskripsi1 = Label(Tumurun, text=deskripsi1, bg="#354f00", fg="white")
    label_deskripsi1.pack(pady=10)
    label_deskripsi2 = Label(Tumurun, text=deskripsi2, bg="#354f00", fg="white")
    label_deskripsi2.pack(pady=3.5)
    label_deskripsi3 = Label(Tumurun, text=deskripsi3, bg="#354f00", fg="white")
    label_deskripsi3.pack(pady=3.5)
    
    Button(window3, width=20, height=1, text="Kembali", fg="white", bg="#57a1f8", command=halaman_utama).place(x=400, y=750)
    Button(window3, width=20, height=1, text="Pesan Tiket", fg="white", bg="#57a1f8", command=Memilih_Wisata).place(x=600, y=750)
    
    def bg_BR(event=None):
        width_BR = window3.winfo_width()
        height_BR = window3.winfo_height()
        resized_image_baru = BG_image.resize((width_BR, height_BR), Image.Resampling.LANCZOS)
        photo_BR = ImageTk.PhotoImage(resized_image_baru)
        BG_label.config(image=photo_BR)
        BG_label.image = photo_BR

    window3.bind("<Configure>", bg_BR)
    window3.mainloop()

def Back_Menu():
    window3.destroy()
    root.deiconify()
    
def Resche_Tiket():
    window2.withdraw()
    global window3
    window3 = tk.Toplevel(window2)
    window3.title("RESCHEDULE TIKET")
    window3.geometry("1200x800")
    window3.configure(bg="White")
    window3.resizable(True, True)

    BG_Image = Image.open(r"C:\prokommmm\Tugas-Besar\Reschedule.png")
    BG_Image = BG_Image.resize((1200, 800), Image.Resampling.LANCZOS)
    Photo_BR = ImageTk.PhotoImage(BG_Image)

    BG_Label = tk.Label(window3, image=Photo_BR)
    BG_Label.place(x=0, y=0, relwidth=1, relheight=1)
    BG_Label.image = Photo_BR

    Lanjut = Frame(window3, width=38, height=1, bg="#354f00")
    Lanjut.place(x=180, y=650)
    Button(Lanjut, width=38, height=1, text="Lanjutkan", fg="white", bg="#57a1f8", command=Memilih_Wisata).pack(pady=0)
    
    Kembali = Frame(window3, width=38, height=1, bg="#354f00")
    Kembali.place(x=180, y=680)
    Button(Kembali, width=38, height=1, text="Kembali", fg="white", bg="#57a1f8", command=halaman_utama).pack(pady=0)

    def bg_BR(event=None):
        Width_BR = window3.winfo_width()
        Height_BR = window3.winfo_height()
        resized_image_baru = BG_Image.resize((Width_BR, Height_BR), Image.Resampling.LANCZOS)
        photo_BR = ImageTk.PhotoImage(resized_image_baru)
        BG_Label.config(image=photo_BR)
        BG_Label.image = photo_BR

    window3.bind("<Configure>", bg_BR)
    window3.mainloop()
    
def Kembali_Menu():
    window3.destroy()
    root.deiconify()

def Kembalikan_Tiket():
    window2.withdraw()
    global window3
    window3 = tk.Toplevel(window2)
    window3.title("REFUND TIKET")
    window3.geometry("1200x800")
    window3.configure(bg="White")
    window3.resizable(True, True)

    BG_imagE = Image.open(r"CC:\prokommmm\Tugas-Besar\Refund.png")
    BG_imagE = BG_imagE.resize((1200, 800), Image.Resampling.LANCZOS)
    pHoto_BR = ImageTk.PhotoImage(BG_imagE)

    Bg_label = tk.Label(window3, image=pHoto_BR)
    Bg_label.place(x=0, y=0, relwidth=1, relheight=1)
    Bg_label.image = pHoto_BR

    Lanjutkan = Frame(window3, width=38, height=1, bg="#354f00")
    Lanjutkan.place(x=180, y=650)
    Button(Lanjutkan, width=38, height=1, text="Lanjutkan", fg="white", bg="#57a1f8", command=Memilih_Wisata).pack(pady=0)
    
    Balek = Frame(window3, width=38, height=1, bg="#354f00")
    Balek.place(x=180, y=680)
    Button(Balek, width=38, height=1, text="Kembali", fg="white", bg="#57a1f8", command=halaman_utama).pack(pady=0)

    def back_BR(event=None):
        width_BarU = window3.winfo_width()
        height_BarU = window3.winfo_height()
        resized_image_baru = BG_imagE.resize((width_BarU, height_BarU), Image.Resampling.LANCZOS)
        photo_BR = ImageTk.PhotoImage(resized_image_baru)
        Bg_label.config(image=photo_BR)
        Bg_label.image = photo_BR

    window3.bind("<Configure>", back_BR)
    window3.mainloop()

def Kembali_Menu():
    window3.destroy()
    root.deiconify()

def Riwayat_Tiket():
    window2.withdraw()
    global window3
    window3 = tk.Toplevel(window2)
    window3.title("RIWAYAT PEMBELIAN")
    window3.geometry("1200x800")
    window3.configure(bg="White")
    window3.resizable(True, True)

    Back_imagE = Image.open(r"C:\prokommmm\Tugas-Besar\Logo_RP.png")
    Back_imagE = Back_imagE.resize((1200, 800), Image.Resampling.LANCZOS)
    pHoto_BR = ImageTk.PhotoImage(Back_imagE)

    Back_label = tk.Label(window3, image=pHoto_BR)
    Back_label.place(x=0, y=0, relwidth=1, relheight=1)
    Back_label.image = pHoto_BR

    def back_BR(event=None):
        width_BarU = window3.winfo_width()
        height_BarU = window3.winfo_height()
        resized_image_baru = Back_imagE.resize((width_BarU, height_BarU), Image.Resampling.LANCZOS)
        photo_BR = ImageTk.PhotoImage(resized_image_baru)
        Back_label.config(image=photo_BR)
        Back_label.image = pHoto_BR

    window3.bind("<Configure>", back_BR)
    window3.mainloop()

# Main window
root = tk.Tk()
root.title("Login and Registration System")
root.geometry("1200x800")

# Load and set background image
bg_image = Image.open(r"C:\prokommmm\Tugas-Besar\wisata.png")
bg_image = bg_image.resize((1200, 800), Image.Resampling.LANCZOS)
bg_photo = ImageTk.PhotoImage(bg_image)

# Create a label with the background image
background_label = tk.Label(root, image=bg_photo)
background_label.place(x=0, y=0, relwidth=1, relheight=1)
background_label.image = bg_photo

# Register Frame
register_frame = Frame(root, bg="white")
Label(register_frame, text="Register", font=("Helvetica", 14), bg="white").grid(row=0, columnspan=2, pady=10)
Label(register_frame, text="Username", bg="white").grid(row=1, column=0, pady=5)
Label(register_frame, text="Password", bg="white").grid(row=2, column=0, pady=5)

entry_register_username = Entry(register_frame)
entry_register_password = Entry(register_frame, show='*')
entry_register_username.grid(row=1, column=1, pady=5)
entry_register_password.grid(row=2, column=1, pady=5)

Button(register_frame, text="Register", command=register).grid(row=3, columnspan=2, pady=10)
Button(register_frame, text="Go to Login", command=show_login_frame).grid(row=4, columnspan=2)

# Login Frame
login_frame = Frame(root, bg="white")
Label(login_frame, text="LOGIN", font=("Helvetica", 14), bg="white").grid(row=0, columnspan=2, pady=10)
Label(login_frame, text="Username", bg="white").grid(row=1, column=0, pady=5)
Label(login_frame, text="Password", bg="white").grid(row=2, column=0, pady=5)

entry_login_username = Entry(login_frame)
entry_login_password = Entry(login_frame, show='*')
entry_login_username.grid(row=1, column=1, pady=5)
entry_login_password.grid(row=2, column=1, pady=5)

Button(login_frame, text="Login", command=login).grid(row=3, columnspan=2, pady=10)
Button(login_frame, text="Go to Register", command=show_register_frame).grid(row=4, columnspan=2)

# Main Frame
main_frame = tk.Frame(root, bg="white")
welcome_label = tk.Label(main_frame, text="Welcome to the Main Page", font=("Helvetica", 14), bg="white")
welcome_label.pack(pady=20)

info_button = tk.Button(main_frame, text="Lanjutkan", command=show_main, font=("Helvetica", 14), bg="white")
info_button.pack(pady=20)

logout_button = tk.Button(main_frame, text="Logout", command=show_login_frame, font=("Helvetica", 14), bg="red", fg="white")
logout_button.pack(pady=20)

# Initially show login frame
show_login_frame()

root.bind("<Configure>", update_background)

root.mainloop()